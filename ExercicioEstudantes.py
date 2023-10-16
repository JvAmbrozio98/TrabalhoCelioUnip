import tkinter.ttk
from tkinter import *
from datetime import date
from tkinter import  filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from ExercicioEstudantesFuncoes import gerarArquivos
from ExercicioEstudantesControle import cursos,listarNomesDeCurso,listarDisciplinas,Aluno,Curso,totalAlunos
background = "#06238D"
frameBG = "#EDEDED"
frameFG = "#06283"

def salvar ():
    A1 = ra.get()
    B1 = nome.get()
    C1 = cpf.get()
    D1 = curso.get()
    E1 = disciplinas.get()

    if (A1 == " " or B1 == "" or C1 == " " or D1 == " " or E1 == " "):
        messagebox.showerror("Erro","Necessario preencher todos os campos")
    else:
        file = openpyxl.load_workbook("InformacoesEstudantes.xlsx")
        sheet = file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=A1)
        sheet.cell(column=2, row=sheet.max_row , value=B1)
        sheet.cell(column=3, row=sheet.max_row , value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=D1)
        sheet.cell(column=5, row=sheet.max_row, value=E1)
        file.save("InformacoesEstudantes.xlsx")
    try:
        img.write("Images/" +str(A1)+".jpg")

    except:
        messagebox.showinfo("Info","Imagem nao encontrada")
    messagebox.showinfo("info","Registrado")
    novaImg = "Images/" + str(A1) + ".jpg"
    alunoAtual = Aluno(A1,B1,C1,D1)
    alunoAtual.disciplina = disciplinas.get()
    alunoAtual.imagem = novaImg
    totalAlunos.append(alunoAtual)
    print(totalAlunos)
    reset()
def serarching():
    text = Search.get()
    text.strip()
    reset()
    for student in totalAlunos:
        busca = str(student.ra)
        busca.strip()
        if text == busca:
            print("Deu")
            nome_entry.insert(0,student.nome)
            ra_entry.insert(0,student.ra)
            cpf_entry.insert(0,student.cpf)
            curso_entry.set(student.id_curso)
            disciplinas_entry.set(student.disciplina)
            novaImagem = PhotoImage(file=student.imagem)
            lbl["image"] = novaImagem
        else:
            messagebox.showerror("ERROR 404","O aluno nao foi encontrado")





def abrir_nova_janela():
    nova_janela = Toplevel(root)
    nova_janela.geometry("1280x600")
    nova_janela.title("Lista de Cursos e Disciplinas")

    tabela = tkinter.ttk.Treeview(nova_janela, columns=("ID", "Nome", "Duração","Disciplinas"), show="headings")
    tabela.heading("ID", text="ID")
    tabela.heading("Nome", text="Nome")
    tabela.heading("Duração", text="Duração")
    tabela.heading("Disciplinas", text="Disciplinas")

    for curso in cursos:
        disclinaDoCurso = curso.listarDisciplinas()
        tabela.insert("", "end", values=(curso.id_curso, curso.nome, curso.duracao,disclinaDoCurso))

    tabela.grid(row=0,column =0)
    pesquisaCuso = StringVar()
    Button(nova_janela,text="Excluir curso").grid(row=1,column =0)
    Label(nova_janela,textvariable=pesquisaCuso,font='arial 13').grid(row=1,column=1)


def reset():
    curso.set('')
    cpf.set('')
    nome.set('')
    disciplinas.set('')
    ra.set('')
    lbl['image'] = img


def mudarFoto ():
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),title="Selec image file",filetypes=(("JPG File","*.jpg"),("PNG File","*.png"),("All files","*")))
    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2
def get_selected_institution(event):
    selected_institution = curso.get()
    update_course_list(selected_institution)


def update_course_list(selected_institution):
    for contador  in cursos:
        if(contador.nome == selected_institution):
            disciplinas_entry["value"] = contador.listarDisciplinas()



def exit ():
    root.destroy()





root = Tk()
root.title("Registro escolar")
root.geometry("1250x700+210+100")
root.config(bg=background)


gerarArquivos()

Label(root,text="Joao Victor Ambrozio RA NC 0666258",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="CADASTRO DE ALUNOS ",width=10,height=2,bg='yellow',font='arial 22 bold').pack(side=TOP,fill=X)

#Botao de pesquisar
Search = StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font='arial 20').place(x=820,y =70)
Srch = Button (root,text="Procurar",compound=LEFT,width=123,font='arial 20 bold',command=serarching)
Srch.place(x = 1060,y = 66)

#Botao de atualizar
updateButton = Button(root,text="Atualizar dados ")
updateButton.place(x=110,y=80)


#Labels para informcoes
Label(root,text="RA",font='arial 13',fg=frameBG,bg=background).place(x=30,y=150)
Label(root,text="NOME",font='arial 13',fg=frameBG,bg=background).place(x=30,y=200)
Label(root,text="CPF",font='arial 13',fg=frameBG,bg=background).place(x=30,y=250)
Label(root,text="CURSO",font='arial 13',fg=frameBG,bg=background).place(x=30,y=300)
Label(root,text="DISCIPLINAS",font='arial 13',fg=frameBG,bg=background).place(x=30,y=350)


#Campos de Preenchimentos
ra = StringVar()
ra_entry = Entry(root,textvariable=ra,font='arial 10')
ra_entry.place(x=150,y=150)

nome = StringVar()
nome_entry = Entry(root,textvariable=nome,font='arial 10')
nome_entry.place(x=150,y=200)

cpf = StringVar()
cpf_entry = Entry(root,textvariable=cpf,font='arial 10')
cpf_entry.place(x=150,y=250)




curso = StringVar()
curso_entry = tkinter.ttk.Combobox(root, values=(listarNomesDeCurso(cursos)), font='arial 12', width=70, state='r',textvariable=curso)
curso_entry.place(x=150,y= 300)
curso_entry.bind("<<ComboboxSelected>>", get_selected_institution)
curso_entry.bind("<<ComboboxSelected>>", update_course_list(get_selected_institution(curso)))



disciplinas = StringVar()
disciplinas_entry = tkinter.ttk.Combobox(root, font='arial 12', width=70, state='r',textvariable=disciplinas)
disciplinas_entry.place(x=150,y= 350)

#Menu lateral

f = Frame(root,bd=3,bg='black',width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)
img = PhotoImage(file='Images/upload photo.png')
lbl = Label(f,bg='black',image=img)
lbl.place(x=0,y=0)
#Botoes do menu lateral
Button(root,text="Mudar foto",width=19,height=2,font="arial 12 bold",bg='lightblue',command=mudarFoto).place(x=1000,y=370)
Button(root,text="Salvar",width=19,height=2,font="arial 12 bold",bg='lightblue',command=salvar).place(x=1000,y=420)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg='lightblue',command=reset).place(x=1000,y=470)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg='lightblue',command=exit).place(x=1000,y=520)
Button(root,text="Editar cursos",width=19,height=2,font="arial 12 bold",bg='lightblue',command=abrir_nova_janela).place(x=1000,y=570)

root.mainloop()